#Employee Record System
import tkinter
from tkinter import*
from tkinter import messagebox
from openpyxl import load_workbook
import pandas as pd

root=Tk() #Main window 
f=Frame(root)
frame1=Frame(root)
frame2=Frame(root)
frame3=Frame(root)
root.title(" Student Management System")
root.geometry("830x395")
root.configure(background="Black")
scrollbar=Scrollbar(root)
scrollbar.pack(side=RIGHT, fill=Y)
firstname=StringVar() #Declaration of all variables
lastname=StringVar()
id=StringVar()
dept=StringVar()
designation=StringVar()
remove_firstname=StringVar()
remove_lastname=StringVar()
searchfirstname=StringVar()
searchlastname=StringVar()
sheet_data=[]
row_data=[]

def emp_dict(*args): #To add a new entry and check if entry already exist in excel sheet
    workbook_name = "sample.xlsx"
    workbook = pd.read_excel(workbook_name)
    wb = load_workbook(workbook_name)
    page = wb.active
    p = 0
    for i in range(len(workbook)):
        row = list(workbook.iloc[i])
        print(row)
        sheet_data.append([])
        sheet_data[p] = row
        p += 1
    print(sheet_data)
    fl = firstname.get()
    fsl = fl.lower()
    ll = lastname.get()
    lsl = ll.lower()
    if (fsl and lsl) in sheet_data:
        print("found")
        messagebox.showerror("Error","This Student already exists")
    else:
        print("not found")
        for info in args:
            page.append(info)
        messagebox.showinfo("Done","Successfully added the student record")
    wb.save(filename=workbook_name)

def add_entries():
    workbook_name = "sample.xlsx"
    wb = load_workbook(workbook_name)
    page = wb.active
    f = firstname.get()
    f1 = f.lower()
    l = lastname.get()
    l1 = l.lower()
    d = dept.get()
    d1 = d.lower()
    de = designation.get()
    de1 = de.lower()
    found = False
    for row in page.iter_rows(min_row=1, max_col=page.max_column, values_only=True):
        if f1 == str(row[0]).lower() and l1 == str(row[1]).lower():
            found = True
            messagebox.showerror("Error","This Student already exists")
            break
    if not found:
        list1=[f1,l1,d1,de1]
        page.insert_rows(1)
        for i, value in enumerate(list1):
            page.cell(row=1, column=i+1, value=value.capitalize())
        messagebox.showinfo("Done","Successfully added the student record")
        wb.save(filename=workbook_name)
        
def add_info(): #for taking user input to add the entries
    frame2.pack_forget()
    frame3.pack_forget()
    emp_first_name=Label(frame1,text="Enter first name of the student: ",bg="red",fg="white")
    emp_first_name.grid(row=1,column=1,padx=10)
    e1=Entry(frame1,textvariable=firstname)
    e1.grid(row=1,column=2,padx=10)
    e1.focus()
    emp_last_name=Label(frame1,text="Enter last name of the student: ",bg="red",fg="white")
    emp_last_name.grid(row=2,column=1,padx=10)
    e2=Entry(frame1,textvariable=lastname)
    e2.grid(row=2,column=2,padx=10)
    emp_dept=Label(frame1,text="Select department of student: ",bg="red",fg="white")
    emp_dept.grid(row=3,column=1,padx=10)
    dept.set("Select Option")
    e4=OptionMenu(frame1,dept,"Select Branch","CSE","ISE","ME","CIV","ECE","EEE")
    e4.grid(row=3,column=2,padx=10)
    emp_desig=Label(frame1,text="Select Semester of Student: ",bg="red",fg="white")
    emp_desig.grid(row=4,column=1,padx=10)
    designation.set("Select Option")
    e5=OptionMenu(frame1,designation,"Select Semester","1","2","3","4","5","6","7","8")
    e5.grid(row=4,column=2,padx=10)
    button4=Button(frame1,text="Add entries",command=add_entries)
    button4.grid(row=5,column=2,pady=10)
    frame1.configure(background="Red")
    frame1.pack(pady=10)

def clear_all(): #for clearing the entry widgets
    frame1.pack_forget
def remove_entries():
    workbook_name = "sample.xlsx"
    wb = load_workbook(workbook_name)
    page = wb.active
    found = False
    for r in range(1, page.max_row + 1):
        if (remove_firstname.get().lower() == str(page.cell(row=r, column=1).value).lower()) and \
                (remove_lastname.get().lower() == str(page.cell(row=r, column=2).value).lower()):
            found = True
            page.delete_rows(r, 1)
            break
    if found:
        wb.save(filename=workbook_name)
        messagebox.showinfo("Done", "Successfully removed the record!")
    else:
        messagebox.showerror("Error", "Student record not found.")


def remove_info(): #for taking user input to remove the entries
    frame1.pack_forget()
    frame3.pack_forget()
    remove_emp_first_name=Label(frame2,text="Enter first name of the Student: ",bg="red",fg="white")
    remove_emp_first_name.grid(row=1,column=1,padx=10)
    e1=Entry(frame2,textvariable=remove_firstname)
    e1.grid(row=1,column=2,padx=10)
    e1.focus()
    remove_emp_last_name=Label(frame2,text="Enter last name of the Student: ",bg="red",fg="white")
    remove_emp_last_name.grid(row=2,column=1,padx=10)
    e2=Entry(frame2,textvariable=remove_lastname)
    e2.grid(row=2,column=2,padx=10)
    button7=Button(frame2,text="Remove entries",command=remove_entries)
    button7.grid(row=3,column=2,pady=10)
    frame2.configure(background="red")
    frame2.pack(pady=10)

def search_entries():
    workbook_name = "sample.xlsx"
    wb = load_workbook(workbook_name)
    page = wb.active
    found = False
    for r in range(1, page.max_row + 1):
        print("Checking row:", r)
        if (searchfirstname.get().lower() == str(page.cell(row=r, column=1).value).lower()) and \
                (searchlastname.get().lower() == str(page.cell(row=r, column=2).value).lower()):
            found = True
            messagebox.showinfo("Student Found", "First Name: " + str(page.cell(row=r, column=1).value) +
                                "\nLast Name: " + str(page.cell(row=r, column=2).value) + "\nDepartment: " +
                                str(page.cell(row=r, column=3).value) + "\nSemester: " +
                                str(page.cell(row=r, column=4).value))
    if found:
        print("Record found")
    else:
        print("Record not found")
        messagebox.showerror("Error", "Student not found.")


def search_info(): #for taking user input to search the entries
    frame1.pack_forget()
    frame2.pack_forget()
    search_emp_first_name=Label(frame3,text="Enter first name of the student: ",bg="red",fg="white")
    search_emp_first_name.grid(row=1,column=1,padx=10)
    e1=Entry(frame3,textvariable=searchfirstname)
    e1.grid(row=1,column=2,padx=10)
    e1.focus()
    search_emp_last_name=Label(frame3,text="Enter last name of the student: ",bg="red",fg="white")
    search_emp_last_name.grid(row=2,column=1,padx=10)
    e2=Entry(frame3,textvariable=searchlastname)
    e2.grid(row=2,column=2,padx=10)
    button9=Button(frame3,text="Search entries",command=search_entries)
    button9.grid(row=3,column=2,pady=10)
    frame3.configure(background="red")
    frame3.pack(pady=10)

def display_emp_record(): #To display all the Student records
    workbook_name = "sample.xlsx"
    wb = load_workbook(workbook_name)
    page = wb.active
    display=Tk()
    display.title("Student Records")
    display.geometry("700x300")
    display.configure(background="white")
    rows = page.max_row
    cols = page.max_column
    for i in range(1,rows + 1):
        for j in range(1,cols + 1):
            e = Entry(display, width=20, fg='red')
            e.grid(row=i, column=j)
            e.insert(END, str(page.cell(row=i, column=j).value))
    display.mainloop()
#Main window GUI design
label=Label(root,text="Welcome to Student Management System",font=("Comic Sans MS",22,"bold"),bg="black",fg="white")
label.pack(pady=10)

button1=Button(f,text="Add Student Record",font=("Comic Sans MS",10,"bold"),bg="red",fg="white",command=add_info)
button1.pack(side=LEFT,padx=10,pady=10)

button2=Button(f,text="Remove Student Record",font=("Comic Sans MS",10,"bold"),bg="red",fg="white",command=remove_info)
button2.pack(side=LEFT,padx=10,pady=10)

button3=Button(f,text="Search Student Record",font=("Comic Sans MS",10,"bold"),bg="red",fg="white",command=search_info)
button3.pack(side=LEFT,padx=10,pady=10)

button5=Button(f,text="Display Student Records",font=("Comic Sans MS",10,"bold"),bg="red",fg="white",command=display_emp_record)
button5.pack(side=LEFT,padx=10,pady=10)

button6=Button(f,text="Exit",font=("Comic Sans MS",10,"bold"),bg="red",fg="white",command=root.destroy)
button6.pack(side=LEFT,padx=10,pady=10)

f.pack()

#Main loop
root.mainloop()
